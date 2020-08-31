import openpyxl


class HomePageData():
    test_fillform_data = [{"Firstname":"Chethan" ,"email":"chethan@gmail.com", "gender":"Male"}, {"Firstname":"Bindu" ,"email":"bindu@gmail.com", "gender":"Female"}]

    @staticmethod
    def gettestdata(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook( "D:\\CK\Appium\\exceldemo1.xlsx" )
        sheet = book.active
        for i in range( 1, sheet.max_row + 1 ):
            if sheet.cell( row=i, column=1 ).value == test_case_name:
                for j in range( 2, sheet.max_column + 1 ):
                    Dict[sheet.cell( row=1, column=j ).value] = sheet.cell( row=i, column=j ).value
                # print( sheet.cell( row=i, column=j ).value )
        return [Dict]